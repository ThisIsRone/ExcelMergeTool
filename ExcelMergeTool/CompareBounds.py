"""
* @author Rone Cao
*
* @email 13592468626@163.com
*
* @create date 2020-02-28 19:19:23
"""

import json
from openpyxl.utils import get_column_letter

class CompareBounds:
    
    def __init__(self,min_col = None,min_row = None,max_col = None,max_row = None,key = None):
        self.min_col = min_col
        self.max_col = max_col
        self.min_row = min_row
        self.max_row = max_row
        self.key = key
        #key健是否合并单元格
        self.own_merge_cell = None
    
    @property
    def bounds(self):
        return (self.min_col,self.max_col, self.min_row,self.max_row)

    def UpdateBounds(self,min_col = None,min_row = None,max_col = None,max_row = None):
        self.min_col = min_col
        self.max_col = max_col
        self.min_row = min_row
        self.max_row = max_row
    
    @property
    def coord(self):
        """
        Excel-style representation of the range
        """
        fmt = "{min_col}{min_row}:{max_col}{max_row}"
        if (self.min_col == self.max_col
            and self.min_row == self.max_row):
            fmt = "{min_col}{min_row}"

        return fmt.format(
            min_col=get_column_letter(self.min_col),
            min_row=self.min_row,
            max_col=get_column_letter(self.max_col),
            max_row=self.max_row)

    @property
    def merge_cell(self):
        return self.own_merge_cell

    def SetMergeCellState(self,merge_cell):
        self.own_merge_cell = merge_cell

    def tostring(self):
        info = "key : " + str(self.key)
        info += "min_col: " +  str(self.min_col)
        info += "max_col: " + str(self.max_col)
        info += "min_row: " + str(self.min_row)
        info += "max_row: " + str(self.max_row)
        return info

