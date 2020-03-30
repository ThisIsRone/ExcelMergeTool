from copy import copy 
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import scr.Helper.DebugHelper as DebugHelper

def PrintPercent(current,total):
	if current == total:
		print(("当前进度{}/{}").format(current,total),end='\n')
	else:
		print(("当前进度{}/{}").format(current,total),end='\r')

#复制单元格
def CopyCell(target_cell,source_cell):
	
	if  isinstance(source_cell,MergedCell):
		# 合并单元格得数据是只读得，需要跳过
		return	
	target_cell.value = source_cell.value
	#copy批注
	if source_cell.comment is not None:
		target_cell.comment = source_cell.comment
	#copy格式
	if source_cell.has_style:	
		target_cell.font = copy(source_cell.font)
		target_cell.border = copy(source_cell.border)
		target_cell.fill = copy(source_cell.fill)
		target_cell.number_format = copy(source_cell.number_format)
		target_cell.protection = copy(source_cell.protection)
		target_cell.alignment = copy(source_cell.alignment)	

#复制整张sheet
def CopySheet(target_sheet,source_sheet):
	#由于设置合并单元格后 使得某些cell变得只读，所以先处理单元格的内容，再处理合并单元格的内容 
	DebugHelper.Log("复制单元格")
	progress = 0
	total = source_sheet.max_row
	for col_items in source_sheet.iter_rows():
		progress += 1
		PrintPercent(progress,total)
		for cell in col_items:
			#copy行列间距(行距设置有问题)
			# if cell.column == 1:
			# 	target_sheet.row_dimensions[cell.row].height = source_sheet.row_dimensions[cell.row].height
			# if cell.row == 1:
			# 	col_letter = get_column_letter(cell.column)
			# 	target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width
			#copy单元格数据
			source_cell =source_sheet.cell(cell.row,cell.column)
			target_cell = target_sheet.cell(cell.row,cell.column) 
			CopyCell(target_cell,source_cell)
	
	#copy合并单元格
	ranges=source_sheet.merged_cells.ranges
	if len(ranges)>0 :
		DebugHelper.Log("复制合并单元格")
		progress = 0
		total = len(ranges)
		for mergecell in ranges:
			progress += 1
			PrintPercent(progress,total)
			target_sheet.merge_cells(mergecell.coord)
