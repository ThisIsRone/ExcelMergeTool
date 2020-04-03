# -*- coding: utf-8 -*-
"""
Created on Fri Dec 13 15:12:40 2019
openpyxl中 bounds元组定义的顺序是：0-3 min_col, min_row, max_col, max_row
@author: admin
"""

from openpyxl import load_workbook
from .SheetReader import SheetReader
import scr.Helper.DebugHelper as DebugHelper
class ExcelReader:
        
    def __init__(self,path_xlsm,excel_title):
        self.excel_title = excel_title
        self.path_xlsm = path_xlsm
        if path_xlsm is None:
             raise Exception("Invalid ExcelReader path_xlsm!", path_xlsm)
        DebugHelper.Log("【读取资源】 " + path_xlsm + " " + excel_title)
        self.workBook = load_workbook(filename = path_xlsm, read_only=False, keep_vba=True)
        # self.dataonly_workBook = load_workbook(filename = path_xlsm, read_only=False, keep_vba=True,data_only=True)
        DebugHelper.Log("【加载完成】 " + path_xlsm)
        sheet_names = self.workBook.get_sheet_names()
        sheet_reader_dic = dict()
        for name in sheet_names:
            sheet = self.workBook.get_sheet_by_name(name)
            # dataonly_sheet = self.dataonly_workBook.get_sheet_by_name(name)
            reader = SheetReader(sheet,None,excel_title)
            sheet_reader_dic[name] =  reader
        self.sheet_reader_dic = sheet_reader_dic
        self.sheet_names = set(sheet_names)
        self.has_diff = False
        DebugHelper.Log("【读取完成】 " + path_xlsm)

    def ApplyDiff2MergeSheet(self,mine_excel_reader,their_excel_reader,sheet_names):
        DebugHelper.Log("【开始合并】 ",self.excel_title)
        for name in sheet_names:
            sheet_reader = self.sheet_reader_dic[name]
            if sheet_reader == None:
                continue
            self.has_diff = True
            DebugHelper.Log("【执行合并】 ",name)
            #将两个字典有差异的内容合并
            mine_key_diffs = mine_excel_reader.sheet_reader_dic[name].key_diffs
            their_key_diffs = their_excel_reader.sheet_reader_dic[name].key_diffs
            if mine_key_diffs and their_key_diffs:
                sheet_reader.ApplyDiff2BodyBounds(mine_key_diffs,their_key_diffs)
        DebugHelper.Log("【完成合并】")

    def OnRelease(self):
        # self.dataonly_workBook.close()
        if self.has_diff:            
            self.workBook.save(self.path_xlsm)
        self.workBook.close()
        self.workBook = None
        self.path_xlsm = None
        self.sheet_names = None
        self.sheet_reader_dic = None
        